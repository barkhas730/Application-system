"""
Тайлангийн хуудасны view функцүүд.

Queryset бэлдэх, статистик тооцоолох ажлыг reports.services гүйцэтгэнэ.
Энд зөвхөн HTTP request/response, Excel export зохицуулна.
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone

from . import services as report_service


def reports_required(view_func):
    """Тайлан харах эрхтэй хэрэглэгчид л зориулсан декоратор."""
    from functools import wraps

    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.can_view_reports:
            messages.error(request, 'Энэ хуудсанд хандах эрх байхгүй.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)

    return wrapper


@login_required
@reports_required
def reports_view(request):
    """
    Тайлангийн үндсэн хуудас.
    Огноо, хэлтэсний шүүлт болон Chart.js графикийн өгөгдлийг бэлдэнэ.
    """
    from accounts.models import CustomUser

    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    dept      = request.GET.get('dept', '')

    # Service layer — role-based queryset болон шүүлт
    apps = report_service.get_report_queryset(
        user=request.user, date_from=date_from, date_to=date_to, dept=dept
    )

    # Статистик тооцоолно
    stats      = report_service.get_report_stats(apps)
    chart_data = report_service.get_chart_data(apps)

    # Хэлтэсний жагсаалт (шүүлтийн select-т ашиглана)
    departments = (
        CustomUser.objects.filter(is_active=True)
        .exclude(department='')
        .values_list('department', flat=True)
        .distinct().order_by('department')
    )

    context = {
        **stats,
        **chart_data,
        'departments': departments,
        'date_from':   date_from,
        'date_to':     date_to,
        'dept':        dept,
    }
    return render(request, 'reports/index.html', context)


@login_required
@reports_required
def reports_export_view(request):
    """
    Тайлангийн өгөгдлийг Excel (.xlsx) файл болгон татаж авна.
    Тайлангийн хуудастай ижил шүүлт хэрэглэнэ.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'openpyxl суулгаагүй байна.')
        return redirect('reports')

    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    dept      = request.GET.get('dept', '')

    # Тайлангийн хуудастай ижил role-based queryset ашиглана
    apps = report_service.get_report_queryset(
        user=request.user, date_from=date_from, date_to=date_to, dept=dept
    ).order_by('-created_at')

    # Excel workbook үүсгэнэ
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Өргөдлүүд'

    # Толгой мөрийн форматлал
    headers     = ['Дугаар', 'Гарчиг', 'Төрөл', 'Ажилтан', 'Хэлтэс', 'Төлөв', 'Ач холбогдол', 'Огноо', 'Дуусах']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(color='FFFFFF', bold=True)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Өгөгдлийн мөрүүд
    for row, app in enumerate(apps, 2):
        ws.cell(row=row, column=1, value=app.app_number)
        ws.cell(row=row, column=2, value=app.title)
        ws.cell(row=row, column=3, value=app.app_type.name if app.app_type else '')
        ws.cell(row=row, column=4, value=app.user.get_full_name() or app.user.username)
        ws.cell(row=row, column=5, value=app.user.department or '')
        ws.cell(row=row, column=6, value=report_service.STATUS_MAP_EXCEL.get(app.status, app.status))
        ws.cell(row=row, column=7, value=report_service.PRIORITY_MAP_EXCEL.get(app.priority, app.priority))
        ws.cell(row=row, column=8, value=app.created_at.strftime('%Y-%m-%d') if app.created_at else '')
        ws.cell(row=row, column=9, value=str(app.due_date) if app.due_date else '')

    # Баганын өргөнийг агуулгад тохируулна
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Файлын нэр — шүүлт байвал "_filtered" нэмнэ
    suffix = timezone.now().strftime('%Y%m%d')
    if date_from or date_to or dept:
        suffix += '_filtered'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="report_{suffix}.xlsx"'
    wb.save(response)
    return response
